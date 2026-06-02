#!/usr/bin/env python3
"""Build a Missouri Disability + IME referral/TPA outreach list for Garcia Family Medicine.
Enriched with verified provider-onboarding contacts (researched 2026-06-02)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

NAVY = "1F3864"; ORANGE = "C55A11"; GREY = "F2F2F2"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical="top", wrap_text=True)
center = Alignment(vertical="center", horizontal="center", wrap_text=True)

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = center; cell.border = border

ws = wb.active
ws.title = "MO Disability + IME List"

headers = ["#", "Organization", "Category", "What they administer / refer", "Best GFM fit",
           "Provider / join URL", "Phone", "Email / contact", "How to apply",
           "Missouri relevance", "Priority", "Status / Notes"]
ws.append(headers)

rows = [
    ["VA disability — C&P exam TPAs (contract licensed MO physicians to perform Compensation & Pension / disability exams)"],
    [1, "QTC Medical Services (Leidos QTC Health Services)", "VA C&P exam TPA",
     "Largest VA contractor for Compensation & Pension (disability) exams nationwide.",
     "IME / disability exams for veterans",
     "qtcm.com/providers", "800-682-9701", "Web form: qtcm.com/contact-us",
     "Apply via Providers page; active unrestricted MO license; orientation before first exam.",
     "Serves MO veterans; needs in-state examiners.", "HIGH",
     "Top target. Note: IMX (#13) now rolls up under Leidos QTC."],
    [2, "VES – Veterans Evaluation Services (Maximus)", "VA C&P exam TPA",
     "Major VA disability exam contractor building a nationwide examiner network.",
     "IME / disability exams for veterans",
     "ves.com/providers", "877-637-8387", "‘Connect With a Recruiter’ form (veteranfeedback@vesservices.com = feedback only)",
     "MO-licensed MD/DO; exams only, no treatment.", "MO examiner network.", "HIGH", "Verified provider page."],
    [3, "Optum Serve (formerly LHI / Logistics Health)", "VA C&P exam TPA",
     "VA & federal medical disability/readiness exam contractor.",
     "IME / disability exams for veterans",
     "providers.optumserve.com/recruitment/provider-form", "Online form only", "Online form only",
     "Complete the ‘Join the Network’ recruitment form; MO license required.",
     "Active VA exam volume in MO.", "HIGH", "Verified recruitment form URL."],
    [4, "Loyal Source Government Services", "VA / federal medical staffing",
     "Staffs clinicians for VA & federal disability/medical exam contracts.",
     "IME / disability exams for veterans",
     "loyalsource.com/technical/mdeproviders", "Form (loyalsource.com/contact-us)", "Web form only",
     "MDE provider-partner inquiry form to perform VA C&P exams at MO-area clinics.",
     "Places clinicians on MO-area contracts.", "MED", "Secondary VA channel."],
    [5, "Sterling Medical Associates", "VA contract exams",
     "Contracts physicians for VA and government medical evaluations (family practice listed).",
     "IME / disability exams for veterans",
     "sterlingmedcorp.com/healthcare-provider-opportunities/physicians", "Web form", "Web form",
     "Apply via Physicians opportunities page for VA-contract placements.",
     "Recruits across states incl. MO.", "MED", "Correct domain: sterlingmedcorp.com (not sterlingmedical.com)."],

    ["Social Security disability — Consultative Exam (CE) channels (SSA pays providers to examine claimants)"],
    [6, "Missouri Disability Determination Services (DDS)", "State SSA disability agency (under DESE Voc Rehab)",
     "Decides MO Social Security disability claims; contracts local providers for Consultative Exams (CEs).",
     "Disability consultative exams (physical & mental)",
     "dese.mo.gov/adult-learning-rehabilitation-services/disability-determination",
     "573-290-5710 (PRO — verify)", "Via PRO phone; SSA PRO locator: ssa.gov/disability/professionals/procontacts.htm",
     "Contact the DDS Professional Relations Officer to enroll as a CE provider; MO license required.",
     "Primary in-state SSA disability channel.", "HIGH",
     "Correction: MO DDS is under DESE (Voc Rehab), NOT DSS/Family Support Division."],
    [7, "MDSI – Medical Diagnostic Services Inc.", "SSA / disability eval vendor",
     "Contracts physicians as independent consultants for disability evaluations nationwide.",
     "Disability consultative exams",
     "mdsiphysicians.com (‘Join Our Network’)", "800-548-9092 / 801-627-1093", "info@mdsiphysicians.com",
     "Contact corporate to join the consultant network.", "Places disability eval volume.", "MED",
     "‘mdexams.com’ didn’t resolve — confirm MDSI is the vendor you meant."],
    [8, "Dane Street", "IME + disability review network",
     "IME, peer review, and disability evaluation network for payers & SSA-related work.",
     "IME + disability exams",
     "site.danestreet.com/physician-network", "Request-info form", "marketing@danestreet.com",
     "Submit interest; panel admission after Credentialing Committee review (ABMS board cert).",
     "Recruits MO examiners.", "HIGH", "Covers both IME and disability — strong fit."],

    ["Workers' comp / auto / legal IME networks (place IME cases with MO examining physicians)"],
    [9, "ExamWorks", "IME network (largest)",
     "Largest IME company in the U.S.; workers' comp, auto, liability, disability IMEs.",
     "IME (workers' comp, auto, legal)",
     "examworks.com/physicians/medical-panel", "206-200-7171 (recruiter — verify)",
     "Erica.seversen@examworks.com (verify)",
     "Join medical panel; most states require ABMS/AOA board cert + ~8 hrs/wk patient care.",
     "Active MO case volume.", "HIGH", "Anchor IME target."],
    [10, "MES Solutions (ExamWorks group)", "IME network",
     "IME and peer review services for insurers and employers.",
     "IME (workers' comp, disability)",
     "messolutions.com/physician-consulting", "Web form", "Web form",
     "Apply via Physician Consulting page; board-certified, active-practice.",
     "Serves MO payers.", "MED", "Same group as ExamWorks — one outreach may cover both."],
    [11, "MCN – Medical Consultants Network", "IME network",
     "IME, peer review, and utilization review provider network.",
     "IME (workers' comp, auto, legal)",
     "mcn.com/join", "Web form", "ntamashiro@mcn.com (email CV)",
     "Email CV or use Join form; credentialing docs at mcn.com/providerdocs.",
     "Recruits MO physicians.", "MED", "Verified."],
    [12, "Genex Services (Enlyte)", "Workers' comp managed care / IME",
     "Workers' comp case management, IME and peer review.",
     "IME (workers' comp)",
     "genexservices.com/solutions/clinical-services/ime/physician-panel", "Web form", "Web form",
     "Apply via IME Physician Panel page (URAC-credentialed).",
     "MO workers' comp footprint.", "MED", "Part of Enlyte/Mitchell."],
    [13, "IMX Medical Mgmt → Leidos QTC Health Services", "IME network",
     "Independent medical exams and record reviews (rebranded under Leidos QTC, Jan 2025).",
     "IME (auto, workers' comp, legal)",
     "qtcm.com/imxmed/physician-panel", "Web form", "Web form",
     "Apply via IMX Physician Panel page; board-certified physicians.",
     "Regional incl. MO.", "LOW", "Now overlaps with #1 (QTC) — apply once."],
    [14, "CorVel", "Workers' comp TPA",
     "Workers' comp third-party administrator; uses IME/exam providers (PPO).",
     "IME (workers' comp)",
     "corvel.com/provider-relations", "Local CorVel rep", "Via Provider Relations page",
     "Download practitioner roster template, return to CorVel network contact; re-credential every 3 yrs.",
     "TPA active in MO.", "MED", "True 'TPA' in the workers'-comp sense."],

    ["Missouri state & government bodies (credentialing / listings / referral)"],
    [15, "Missouri Division of Workers' Compensation (DOLIR)", "State agency",
     "Administers MO workers' comp; physician IMEs (§287.140) & disability ratings.",
     "IME / impairment ratings",
     "labor.mo.gov/dwc/healthcare-providers", "573-751-4231 / 800-775-2667", "workerscomp@labor.mo.gov",
     "Informational only — no network. Physician of claimant's choosing is permitted.",
     "Statewide authority on MO workers' comp IMEs.", "LOW",
     "Use to understand MO rating rules (AMA Guides) before marketing — not an enrollment channel."],
    [16, "Missouri Veterans Commission", "State veterans agency",
     "State veterans services; CVSOs help vets file disability claims (need nexus/exam support).",
     "Nexus letters / veteran disability exams",
     "mvc.dps.mo.gov/service (CVSO locator)", "573-751-3779 HQ · 816-356-4075 KC", "mvc.dps.mo.gov/contact-us",
     "Use the Service Officer Locator to reach the nearest County Veterans Service Officer.",
     "Direct line to MO veterans needing evidence.", "HIGH", "High-value referral relationship."],

    ["Referral sources — attorneys & veteran orgs (send claimants who need IMEs / nexus letters)"],
    [17, "Missouri workers'-comp & personal-injury attorneys", "Attorney referral",
     "Plaintiff & defense attorneys routinely order IMEs and medical opinions.",
     "IME / medical opinion letters",
     "Firm sites (e.g., Boyd, Kenter, Thomas & Parrish — already in your files)", "Per firm", "Per firm",
     "Targeted outreach with fee schedule + work sample; you already emailed BKTP.",
     "Local Jackson County / KC metro firms.", "HIGH",
     "Warm channel — you delivered the M. Mendez IME for BKTP."],
    [18, "Missouri Association of Trial Attorneys (MATA)", "Attorney association",
     "Plaintiff trial bar; members need independent medical experts.",
     "IME / expert medical opinions",
     "matanet.org (advertising/sponsorship)", "Site contact", "Contact form",
     "Sponsor / advertise to reach the plaintiff bar; no public expert directory.",
     "Statewide plaintiff attorney audience.", "MED",
     "Correct domain: matanet.org (not matalaw.com)."],
    [19, "DAV · VFW · American Legion (Missouri departments)", "Veteran service orgs",
     "VSOs assist veterans with disability claims; refer for nexus letters & exams.",
     "Nexus letters / veteran disability exams",
     "DAV: davwebsites.dav.org/mo · VFW: vfwmo.org/veteran-service-officer-resources · Legion: missourilegion.org/contact-us",
     "DAV KC 816-765-8787 · VFW HQ 573-636-8761 · Legion Svc 573-761-4143",
     "services@missourilegion.org (Legion); DAV/VFW via forms",
     "Introduce Veterans@Work IME + Nexus services to MO service officers.",
     "Large MO veteran membership.", "HIGH", "Legion domain is missourilegion.org."],
    [20, "County Veterans Service Officers (Jackson, Cass, etc.)", "Local referral",
     "Frontline help for vets filing claims; trusted referrers.",
     "Nexus letters / veteran disability exams",
     "jacksongov.org/Residents/Health-Services/Veteran-BenefitsServices", "816-325-5883 / 816-325-5882", "Call office",
     "Independence office (Mon–Fri 8–4) serves Jackson & Cass Counties.",
     "Local, high-trust referrals.", "HIGH", "Start with your home counties."],
]

ncols = len(headers); r = 2; data_row_count = 0
for row in rows:
    if len(row) == 1:
        ws.cell(row=r, column=1, value=row[0])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=ORANGE)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = border
    else:
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap; cell.border = border
            if c == 11:
                p = str(val).upper()
                color = {"HIGH": "C6EFCE", "MED": "FFEB9C", "LOW": "F2F2F2"}.get(p)
                if color: cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = center
            if c == 1: cell.alignment = center
        data_row_count += 1
    r += 1

style_header(ws, ncols)
widths = [4, 27, 18, 30, 18, 30, 20, 26, 30, 22, 8, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "C2"

# How-to-use sheet
ws2 = wb.create_sheet("How to use")
notes = [
    ["Garcia Family Medicine — Missouri Disability & IME Outreach List", ""],
    ["Contacts researched / verified", "2026-06-02. Websites are official domains; most networks funnel applicants through web forms, so exact recruiter emails/phones are often not published. Items marked 'verify' should be confirmed on the live page before use. No phone/email was invented."],
    ["", ""],
    ["Three ways the work flows in", ""],
    ["1. VA disability TPAs", "QTC, VES, Optum Serve, Loyal Source, Sterling credential you as a network examiner for veteran C&P disability exams. Highest volume."],
    ["2. SSA disability (CE)", "Missouri DDS (under DESE Voc Rehab) pays providers for Consultative Exams. MDSI / Dane Street also place disability evals."],
    ["3. IME networks & attorneys", "ExamWorks, Dane Street, MCN, Genex, CorVel place workers'-comp/auto/legal IMEs. Attorneys & VSOs send claimants directly for IMEs and nexus letters."],
    ["", ""],
    ["Priority key", "HIGH = start here (volume or warm relationship). MED = solid secondary. LOW = verify fit / informational only."],
    ["", ""],
    ["Corrections from research", "• MO DDS is under DESE (Voc Rehab), not DSS/FSD.  • IMX is now Leidos QTC (overlaps #1).  • MATA domain is matanet.org.  • American Legion MO is missourilegion.org.  • Sterling is sterlingmedcorp.com.  • 'mdexams.com' → MDSI (mdsiphysicians.com) — confirm intended vendor."],
    ["", ""],
    ["Suggested first moves", "1) Apply to QTC + VES + Optum Serve (VA). 2) Call Missouri DDS PRO about CE enrollment. 3) Join ExamWorks + Dane Street panels. 4) Re-contact BKTP and pitch CVSOs/DAV/VFW/Legion with the Veterans@Work IME + Nexus packet."],
    ["", ""],
    ["Assets to attach", "• IME service + pricing sheet  • Veterans@Work IME Referral Flyer + sample MoU  • Nexus Letter Services doc  • completed IME report (M. Mendez) as a work sample.  • The outreach packet (5 email templates) accompanies this list."],
]
for row in notes: ws2.append(row)
ws2.column_dimensions["A"].width = 30; ws2.column_dimensions["B"].width = 100
from openpyxl.styles import Font as F
ws2["A1"].font = F(bold=True, size=14, color=NAVY)
for rr in range(2, len(notes) + 1):
    a = ws2.cell(row=rr, column=1)
    a.font = F(bold=True, color=NAVY) if a.value else F()
    a.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.cell(row=rr, column=2).alignment = Alignment(vertical="top", wrap_text=True)

out = "/home/user/voicerx-site/GFM_Missouri_Disability_IME_List.xlsx"
wb.save(out)
print("saved", out, "| data rows:", data_row_count)
